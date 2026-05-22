"""
Sistema de Ficha Médica Ocupacional - CASISO S.A.S.
Dr. Martín Lino Zambrano
Versión Cloud — Render.com + Supabase
"""
import os, io, json
from copy import copy
from datetime import date
from flask import Flask, request, jsonify, send_file, Response
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
PLANTILLAS = os.path.join(BASE, "plantillas")


# ── SUPABASE (base de datos en la nube) ──────────────────────
# En producción (Render) se usan variables de entorno.
# En local funciona con lista en memoria como respaldo.
SUPABASE_URL  = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY  = os.environ.get("SUPABASE_KEY", "")
TABLE_NAME    = "matriz_evaluaciones"

supa = None
if SUPABASE_URL and SUPABASE_KEY:
    try:
        from supabase import create_client
        supa = create_client(SUPABASE_URL, SUPABASE_KEY)
        print("✓ Conectado a Supabase")
    except Exception as e:
        print(f"⚠ Supabase no disponible: {e}")
        supa = None

# Fallback local (cuando no hay Supabase configurado)
_matriz_local = []

def db_insertar(fila):
    if supa:
        try:
            supa.table(TABLE_NAME).insert({"data": fila}).execute()
            return
        except Exception as e:
            print(f"⚠ Supabase insert error: {e}")
    _matriz_local.append(fila)

def db_listar():
    if supa:
        try:
            res = supa.table(TABLE_NAME).select("*").order("id").execute()
            supabase_data = [r["data"] for r in res.data]
            # Combinar con locales si hay
            return supabase_data + _matriz_local
        except Exception as e:
            print(f"⚠ Supabase list error: {e}")
    return _matriz_local

def db_contar():
    if supa:
        try:
            res = supa.table(TABLE_NAME).select("id", count="exact").execute()
            return (res.count or 0) + len(_matriz_local)
        except Exception as e:
            print(f"⚠ Supabase count error: {e}")
    return len(_matriz_local)

def db_limpiar():
    global _matriz_local
    if supa:
        try:
            supa.table(TABLE_NAME).delete().neq("id", 0).execute()
        except Exception as e:
            print(f"⚠ Supabase delete error: {e}")
    _matriz_local = []
# ─────────────────────────────────────────────────────────────

def sw(ws, addr, value):
    """safe_write: escribe en celda maestra, aplica Arial 10, wrap_text y ajusta altura."""
    from openpyxl.utils import coordinate_to_tuple
    from openpyxl.styles import Alignment, Font

    def _write(cell, val):
        cell.value = val
        if val:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if isinstance(val, str):
                cols_aprox = 80
                lineas = val.count('\n') + max(1, len(val) // cols_aprox)
                if lineas > 1:
                    current = ws.row_dimensions[cell.row].height or 15
                    ws.row_dimensions[cell.row].height = max(current, min(lineas * 14, 300))

    try:
        _write(ws[addr], value)
    except AttributeError:
        try:
            row, col = coordinate_to_tuple(addr)
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    _write(ws.cell(mr.min_row, mr.min_col), value)
                    return
        except Exception:
            pass

app = Flask(__name__)

# ─────────────────────────────────────────────
#  LLENADO DE PLANTILLAS
# ─────────────────────────────────────────────
def chk(v): return "X" if v else ""
def val(v): return str(v) if v else ""

def llenar_hoja1(d):
    wb = load_workbook(os.path.join(PLANTILLAS, "PRIMERA_HOJA.xlsx"))
    ws = wb.active

    # A. Datos establecimiento — fila 5 (valores)
    sw(ws, "A5",  d.get("institucion",""))
    sw(ws, "P5",  d.get("ruc",""))
    sw(ws, "Y5",  d.get("ciiu",""))
    sw(ws, "AB5", d.get("establecimiento",""))
    sw(ws, "AO5", d.get("nroHistoria",""))
    sw(ws, "BA5", d.get("nroArchivo",""))

    # A. Nombres — fila 8 (datos). Labels están en filas 6-7 (A6:Q7 merged)
    sw(ws, "A8",  d.get("primerApellido",""))
    sw(ws, "R8",  d.get("segundoApellido",""))
    sw(ws, "AG8", d.get("primerNombre",""))
    sw(ws, "AQ8", d.get("segundoNombre",""))

    # Grupo atención prioritaria — fila 13 (debajo de labels 10-12)
    sw(ws, "A13", "X" if d.get("embarazada")      else "NO")
    sw(ws, "F13", "X" if d.get("discapacidad")    else "NO")
    sw(ws, "J13", "X" if d.get("enfCatastrofica") else "NO")
    sw(ws, "M13", "X" if d.get("adultoMayor")     else "NO")
    sexo = d.get("sexo","")
    if sexo == "Hombre": sw(ws, "T13", "X")
    if sexo == "Mujer":  sw(ws, "W13", "X")

    # Fecha nacimiento — Y13=año, AB13=mes, AD13=día (fila 13 datos)
    fn = d.get("fechaNacimiento","")
    if fn and len(fn) >= 10:
        sw(ws, "Y13",  fn[0:4])
        sw(ws, "AB13", fn[5:7])
        sw(ws, "AD13", fn[8:10])
    sw(ws, "AE13", d.get("edad",""))
    sw(ws, "AG13", d.get("grupoSanguineo",""))
    sw(ws, "AO13", d.get("lateralidad",""))

    # B. Motivo de consulta
    # I16:W16 -> I16 (puesto CIUO dato), X16:AF16 -> LABEL, AG16:BF16 -> fechaAtencion dato
    sw(ws, "I16",  d.get("puestoCIUO",""))
    sw(ws, "AG16", d.get("fechaAtencion",""))
    # A17:M17=label fechaIngreso, A18:AB18=dato; AC17:AJ17=label, AC18:AK18=dato; AL17:BF17=label, AL18:BF18=dato
    sw(ws, "A18",  d.get("fechaIngreso",""))
    sw(ws, "AC18", d.get("fechaReintegro",""))
    sw(ws, "AL18", d.get("fechaUltimoDia",""))

    # Tipo evaluacion — labels en row 20, checkboxes en celdas vacías de row 20
    # A20:E20=INGRESO label, F20:M20=checkbox; N20:X20=PERIÓDICO label, Y20:AB20=checkbox
    # AC20:AG20=REINTEGRO label, AH20:AK20=checkbox; AL20:AY20=RETIRO label, AZ20:BF20=checkbox
    tipo = d.get("tipoEvaluacion","")
    if tipo == "INGRESO":   sw(ws, "F20",  "X")
    if tipo == "PERIÓDICO": sw(ws, "Y20",  "X")
    if tipo == "REINTEGRO": sw(ws, "AH20", "X")
    if tipo == "RETIRO":    sw(ws, "AZ20", "X")
    sw(ws, "A21",  d.get("motivoObs",""))

    # C. Antecedentes personales — A25, A27 (celdas dato)
    texto_ant = d.get("antClinico","")
    sw(ws, "A25", texto_ant)
    ws.row_dimensions[25].height = max(45, (len(texto_ant)//80+1)*15) if texto_ant else 45

    texto_fam = d.get("antFamiliares","")
    sw(ws, "A27", texto_fam)
    ws.row_dimensions[27].height = max(45, (len(texto_fam)//80+1)*15) if texto_fam else 45

    # Transfusiones — row 29=labels, row 30=datos/checkboxes
    # N29:Q29="SI" label → N30:Q30=checkbox; R29:U29="NO" label → R30:U30=checkbox
    transfusion = d.get("autTransfusion","")
    if transfusion == "SI": sw(ws, "N30", "X")
    if transfusion == "NO": sw(ws, "R30", "X")
    # AG29:AJ29="SI" label → AG30:AJ30=checkbox; AK29:BC29="¿Cuál?" label → AK30:BC30=dato
    tto = d.get("ttoHormonal","")
    if tto == "SI": sw(ws, "AG30", "X")
    if tto == "NO": sw(ws, "BD30", "X")
    sw(ws, "AK30", d.get("ttoHormonalCual",""))

    # Gineco obstétricos — rows 32-33=labels, row 34=datos
    # A32:V33=FUM label → A34:V34=dato; W32:Z33=GESTAS label → W34:Z34=dato
    if d.get("ginecoNoAplica"):
        sw(ws, "A34", "NO APLICA")
    else:
        sw(ws, "A34",  d.get("fum",""))
        sw(ws, "W34",  d.get("gestas",""))
        sw(ws, "AA34", d.get("partos",""))
        sw(ws, "AD34", d.get("cesareas",""))
        sw(ws, "AG34", d.get("abortos",""))
        sw(ws, "AM34", d.get("metPlanFem",""))

    # Reproductivos masculinos — A39:M40=label dato, N39:S40=tiempo dato
    # AD40/AO40/AX40 = labels SI/NO/NO RESPONDE → datos en AE41/AO41/AX41
    sw(ws, "A39",  d.get("examRepMasc",""))
    sw(ws, "N39",  d.get("tiempoRepMasc",""))
    metPlan = d.get("metPlanMasc","")
    if metPlan == "SI" or d.get("metPlanMascSi"):
        sw(ws, "AE41", d.get("metPlanMascCual",""))
    if metPlan == "NO":       sw(ws, "AO41", "X")
    if metPlan == "NO_RESP":  sw(ws, "AX41", "X")
    if d.get("metPlanMascCual"): sw(ws, "AE41", d.get("metPlanMascCual",""))

    # Consumo sustancias — filas 45,46,47
    sw(ws, "M45",  d.get("tabacoT",""))
    if d.get("tabacoEx"):  sw(ws, "R45", "X")
    sw(ws, "W45",  d.get("tabacoAbs",""))
    if d.get("tabacoNo"):  sw(ws, "AA45", "X")

    sw(ws, "M46",  d.get("alcoholT",""))
    if d.get("alcoholEx"): sw(ws, "R46", "X")
    sw(ws, "W46",  d.get("alcoholAbs",""))
    if d.get("alcoholNo"): sw(ws, "AA46", "X")

    sw(ws, "M47",  d.get("otraSustanciaT",""))
    if d.get("otraSustanciaEx"): sw(ws, "R47", "X")
    sw(ws, "W47",  d.get("otraSustanciaAbs",""))
    if d.get("otraSustanciaNo"): sw(ws, "AA47", "X")
    sw(ws, "A48",  d.get("condEspecial",""))

    # Estilo de vida — AH45=X actFisica, AK45=tiempo, AT45=cuál
    if d.get("actFisica"): sw(ws, "AH45", "X")
    act_t = d.get("actFisicaT","")
    act_c = d.get("actFisicaCual","")
    sw(ws, "AK45", act_t)
    sw(ws, "AT45", act_c)
    if act_t: ws.row_dimensions[45].height = max(30, (len(act_t)//20+1)*14)

    # Medicación habitual — AT45=cuál, BB45=cantidad
    med_c = d.get("medicHabitualCual","")
    med_q = d.get("medicHabitualCantidad","")
    if d.get("medicHabitual") or med_c:
        sw(ws, "AT45", med_c)
        sw(ws, "BB45", med_q)
        if med_c: ws.row_dimensions[45].height = max(ws.row_dimensions[45].height or 30, (len(med_c)//20+1)*14)

    # D. Enfermedad actual — A51:BF52 (A50 es el header de sección)
    enf = d.get("enfermedadActual","")
    sw(ws, "A51", enf)
    if enf:
        lineas = max(3, len(enf)//100 + enf.count('\n') + 2)
        ws.row_dimensions[51].height = max(30, lineas * 14)
        ws.row_dimensions[52].height = max(30, lineas * 14)

    # E. Constantes vitales — fila 56 (datos), fila 55 son labels
    sw(ws, "A56",  d.get("temperatura",""))
    sw(ws, "H56",  d.get("pa",""))
    sw(ws, "O56",  d.get("fc",""))
    sw(ws, "V56",  d.get("fr",""))
    sw(ws, "AB56", d.get("spo2",""))
    sw(ws, "AF56", d.get("peso",""))
    sw(ws, "AH56", d.get("talla",""))
    sw(ws, "AL56", d.get("imc",""))
    sw(ws, "AR56", d.get("perAbdominal",""))

    # F. Examen físico regional — texto en A72
    regiones = d.get("examenFisicoRegiones", {})
    obs_adicional = d.get("examenFisicoObs","")
    if regiones:
        con_patologia, normales = [], []
        for region, data in regiones.items():
            if isinstance(data, dict):
                if data.get("checked"):
                    h = data.get("hallazgo","").strip()
                    con_patologia.append(f"{region}: {h}" if h else f"{region}: PATOLOGÍA")
                else:
                    normales.append(region.split(".")[0].strip())
        lineas = []
        if con_patologia: lineas.append("CON HALLAZGOS: " + "; ".join(con_patologia))
        if normales:       lineas.append("NORMAL: " + ", ".join(normales))
        if obs_adicional:  lineas.append("Obs: " + obs_adicional)
        texto_examen = "\n".join(lineas) if lineas else "Sin hallazgos patológicos."
    else:
        texto_examen = obs_adicional

    sw(ws, "A72", texto_examen)
    if texto_examen:
        h = max(45, (len(texto_examen)//100 + texto_examen.count('\n') + 2) * 15)
        ws.row_dimensions[72].height = h

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def llenar_hoja2(d):
    wb = load_workbook(os.path.join(PLANTILLAS, "SEGUNDA_HOJA.xlsx"))
    ws = wb.active

    # Puesto de trabajo — A3:F5 -> maestra A3 (bloque blanco arriba de números)
    sw(ws, "A3", d.get("puestoRiesgo",""))

    # Actividades importantes — G5,I5,K5,M5,N5,O5,P5 (fila 5, debajo de los números 1-7)
    actividades = d.get("actividades", ["","","","","","",""])
    for i, col in enumerate(["G5","I5","K5","M5","N5","O5","P5"]):
        if i < len(actividades):
            sw(ws, col, actividades[i])
            if actividades[i]:
                ws.row_dimensions[5].height = max(ws.row_dimensions[5].height or 30,
                    (len(actividades[i])//15+1)*14)

    # Riesgos
    RIESGOS_FILAS = {
        "Temperaturas altas": 6, "Temperaturas bajas": 7,
        "Radiación Ionizante": 8, "Radiación No Ionizante": 9,
        "Ruido": 10, "Vibración": 11, "Iluminación": 12,
        "Ventilación": 13, "Fluido eléctrico": 14, "Otros (Físico)": 15,
        "Falta de señalización, aseo, desorden": 16,
        "Atrapamiento entre Máquinas y o superficies": 17,
        "Atrapamiento entre objetos": 18, "Caída de objetos": 19,
        "Caídas al mismo nivel": 20, "Caídas a diferente nivel": 21,
        "Pinchazos": 22, "Cortes": 23, "Choques /colisión vehicular": 24,
        "Atropellamientos por vehículos": 25, "Proyección de fluidos": 26,
        "Proyección de partículas – fragmentos": 27,
        "Contacto con superficies de trabajos": 28, "Contacto eléctrico": 29,
        "Polvos ": 31, "Sólidos": 32, "Humos": 33, "líquidos ": 34,
        "vapores": 35, "Aerosoles": 36, "Neblinas ": 37, "Gaseosos": 38,
        "Virus ": 40, "Hongos": 41, "Bacterias ": 42, "Parásitos ": 43,
        "Exposición a vectores": 44, "Exposición a animales selváticos ": 45,
        "Manejo manual de cargas": 47, "Movimiento repetitivos": 48,
        "Posturas forzadas": 49, "Trabajos con PVD": 50,
        "Diseño Inadecuado del puesto": 51,
        "Monotonía del trabajo": 53, "Sobrecarga laboral": 54,
        "Minuciosidad de la tarea ": 55, "Alta responsabilidad": 56,
        "Autonomía  en la toma de decisiones": 57,
        "Supervisión y estilos de dirección deficiente": 58,
        "Conflicto de rol": 59, "Falta de Claridad en las funciones": 60,
        "Incorrecta distribución del trabajo ": 61,
        "Turnos rotativos": 62, "Relaciones interpersonales ": 63,
        "inestabilidad laboral": 64, "Amenaza Delincuencial": 65,
    }
    ACT_LETTER = {0:"G",1:"I",2:"K",3:"M",4:"N",5:"O",6:"P"}
    riesgos_data = d.get("riesgos", {})
    for nombre_riesgo, fila in RIESGOS_FILAS.items():
        for key, vals in riesgos_data.items():
            if key.strip().lower() == nombre_riesgo.strip().lower() or \
               key.strip() in nombre_riesgo or nombre_riesgo.strip() in key:
                for i, checked in enumerate(vals[:7]):
                    if checked:
                        sw(ws, f"{ACT_LETTER.get(i,'G')}{fila}", "X")
                break

    # Medidas preventivas — A67:F71 label, G67:H68 dato (2 filas merged)
    # Las filas 69-71 son del label (A67:F71) — deben tener altura mínima
    med = d.get("medidasPreventivas","")
    sw(ws, "G67", med)
    lineas_med = (med.count("\n") + 1) if med else 1
    h_med = max(14, (lineas_med * 14) // 2 + 4)
    ws.row_dimensions[67].height = h_med
    ws.row_dimensions[68].height = h_med
    for r in [69, 70, 71]:
        ws.row_dimensions[r].height = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def llenar_hoja3(d):
    wb = load_workbook(os.path.join(PLANTILLAS, "TERCERA_HOJA.xlsx"))
    ws = wb.active

    # H. Antecedentes laborales (filas 7-25)
    al = d.get("antLaborales", [{}])
    filas_al = [7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25]
    for i, a in enumerate(al[:len(filas_al)]):
        f = filas_al[i]
        sw(ws, f"B{f}",  a.get("centro",""))
        sw(ws, f"J{f}",  a.get("actividades",""))
        sw(ws, f"W{f}",  chk(a.get("anterior")))
        sw(ws, f"Y{f}",  chk(a.get("actual")))
        sw(ws, f"AA{f}", a.get("tiempo",""))
        sw(ws, f"AC{f}", chk(a.get("incidente")))
        sw(ws, f"AE{f}", chk(a.get("accidente")))
        sw(ws, f"AH{f}", chk(a.get("ep")))
        sw(ws, f"AK{f}", chk(a.get("iesssi")))
        sw(ws, f"AM{f}", chk(a.get("iessno")))
        sw(ws, f"AO{f}", a.get("fecha",""))
        sw(ws, f"AR{f}", a.get("especificar",""))

    # I. Extra laborales fila 28-29
    sw(ws, "A28", d.get("actExtraLaborales",""))

    # J. Exámenes filas 36-40 (fila 35 = labels: NOMBRE/FECHA/RESULTADOS)
    examenes = d.get("examenes", [])
    for i, ex in enumerate(examenes[:5]):
        f = 36 + i
        sw(ws, f"B{f}",  ex.get("nombre",""))
        sw(ws, f"M{f}",  ex.get("fecha",""))
        sw(ws, f"T{f}",  ex.get("resultado",""))
    sw(ws, "B41", "OBSERVACIONES: " + d.get("examenesObs",""))

    # K. Diagnósticos filas 45-50
    diagnosticos = d.get("diagnosticos", [])
    for i, dx in enumerate(diagnosticos[:6]):
        f = 45 + i
        sw(ws, f"B{f}",  dx.get("cie10",""))
        sw(ws, f"Q{f}",  dx.get("desc",""))
        pre = dx.get("tipo","") == "PRE"
        def_ = dx.get("tipo","") == "DEF"
        sw(ws, f"BE{f}", "X" if pre  else "")
        sw(ws, f"BJ{f}", "X" if def_ else "")

    # L. Aptitud — P53=APTO, AE53=APTO_OBS, AS53=APTO_LIM, BB53=NO_APTO (celdas checkbox, no labels)
    aptitud = d.get("aptitud","")
    if aptitud == "APTO":     sw(ws, "P53",  "X")
    if aptitud == "APTO_OBS": sw(ws, "AE53", "X")
    if aptitud == "APTO_LIM": sw(ws, "AS53", "X")
    if aptitud == "NO_APTO":  sw(ws, "BB53", "X")
    obs_apt = d.get("aptitudObs","")
    sw(ws, "B54", obs_apt)
    if obs_apt:
        ws.row_dimensions[54].height = max(14, (obs_apt.count("\n")+1)*14)

    # M. Recomendaciones fila 59
    reco3 = d.get("recomendaciones","")
    sw(ws, "B59", reco3)
    if reco3:
        ws.row_dimensions[59].height = max(14, (reco3.count("\n")+1)*14)

    # N. Retiro — AG65=SI checkbox, AW65=NO checkbox (W65/AI65 son labels)
    ret_eval = d.get("retiroEval","")
    if ret_eval == "SI": sw(ws, "AG65", "X")
    if ret_eval == "NO": sw(ws, "AW65", "X")
    ret_rel = d.get("retiroRelacionado","")
    if ret_rel == "SI": sw(ws, "AG66", "X")
    if ret_rel == "NO": sw(ws, "AW66", "X")
    sw(ws, "B68",  d.get("retiroObs",""))  # B67=label "Observación:", B68=dato

    # O. Datos profesional filas 73-74
    sw(ws, "K73",  d.get("nombreProf",""))
    sw(ws, "AB73", d.get("codigoMedico",""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def llenar_certificado(d):
    wb = load_workbook(os.path.join(PLANTILLAS, "CERTIFICADO_DE_APTITUP_LABORAL.xlsx"))
    ws = wb.active

    # A. Datos establecimiento — fila 4 (valores)
    sw(ws, "A4",  d.get("institucion",""))
    sw(ws, "L4",  d.get("ruc",""))
    sw(ws, "R4",  d.get("ciiu",""))
    sw(ws, "V4",  d.get("establecimiento",""))
    sw(ws, "AC4", d.get("nroHistoria",""))
    sw(ws, "AI4", d.get("nroArchivo",""))

    # Nombres — fila 6 (A6, J6, Q6, X6, AD6, AG6)
    sw(ws, "A6",  d.get("primerApellido",""))
    sw(ws, "J6",  d.get("segundoApellido",""))
    sw(ws, "Q6",  d.get("primerNombre",""))
    sw(ws, "X6",  d.get("segundoNombre",""))
    sw(ws, "AD6", d.get("sexo",""))
    sw(ws, "AG6", d.get("puestoCIUO",""))

    # B. Fecha emisión — K10=año, M10=mes, O10=día (labels K11/M11/O11)
    fa = d.get("fechaAtencion","")
    if fa and len(fa) >= 10:
        sw(ws, "K10", fa[0:4])   # año completo en celda merged K10:L10
        sw(ws, "M10", fa[5])     # decena del mes
        sw(ws, "N10", fa[6])     # unidad del mes
        sw(ws, "O10", fa[8])     # decena del día
        sw(ws, "P10", fa[9])     # unidad del día

    # Tipo evaluación — L12/U12/AC12/AI12
    tipo = d.get("tipoEvaluacion","")
    if tipo == "INGRESO":   sw(ws, "L12",  "X")
    if tipo == "PERIÓDICO": sw(ws, "U12",  "X")
    if tipo == "REINTEGRO": sw(ws, "AC12", "X")
    if tipo == "RETIRO":    sw(ws, "AI12", "X")

    # C. Aptitud — I17=checkbox APTO, S17=APTO_OBS, AC17=APTO_LIM, AK17=NO_APTO
    # A17/J17/U17/AE17 son LABELS — no se escriben
    aptitud = d.get("aptitud","")
    if aptitud == "APTO":     sw(ws, "I17",  "X")
    if aptitud == "APTO_OBS": sw(ws, "S17",  "X")
    if aptitud == "APTO_LIM": sw(ws, "AC17", "X")
    if aptitud == "NO_APTO":  sw(ws, "AK17", "X")
    sw(ws, "A19",  d.get("aptitudObs",""))

    # D. Recomendaciones — A24:AL26 son 3 filas merged
    # Altura correcta: cada línea ~14pt, distribuida en 3 filas
    reco = d.get("recomendaciones","")
    sw(ws, "A24", reco)
    lineas_reco = (reco.count("\n") + 1) if reco else 1
    h_por_fila = max(14, (lineas_reco * 14) // 3 + 4)
    for r in [24, 25, 26]:
        ws.row_dimensions[r].height = h_por_fila

    # E. Datos profesional
    # A33:D33 = label "NOMBRE Y APELLIDO" → nombre va en E33:K33
    # L33:N33 = label "CÓDIGO MÉDICO"    → código va en O33:S33
    sw(ws, "E33", d.get("nombreProf",""))
    sw(ws, "O33", d.get("codigoMedico",""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

#  MAPA DE COLORES INDEXED → RGB (paleta Excel)
# ─────────────────────────────────────────────
INDEXED_TO_RGB = {
    9:  'FFFFFFFF',  # Blanco
    27: 'FFCCFFFF',  # Celeste claro
    31: 'FFCCCCFF',  # Lavanda/azul claro
    42: 'FFCCFFCC',  # Verde claro
    0:  'FF000000', 1:  'FFFFFFFF', 2:  'FFFF0000', 3:  'FF00FF00',
    4:  'FF0000FF', 5:  'FFFFFF00', 6:  'FFFF00FF', 7:  'FF00FFFF',
    22: 'FFC0C0C0', 26: 'FFFFFFCC', 43: 'FFFFFF99', 44: 'FF99CCFF',
    45: 'FFFF99CC', 47: 'FFFFCC99', 49: 'FF33CCCC', 51: 'FFFFCC00',
    53: 'FFFF6600', 62: 'FF333399', 63: 'FF000000', 64: 'FFFFFFFF',
}

def get_fill_rgb(fill):
    """Obtiene el color RGB de un fill, convirtiendo indexed si es necesario."""
    if not fill or fill.fill_type != 'solid':
        return None
    fg = fill.fgColor
    if fg.type == 'rgb' and fg.rgb not in ('00000000', 'FF000000'):
        return fg.rgb
    if fg.type == 'indexed':
        return INDEXED_TO_RGB.get(fg.indexed)
    return None

def normalize_border_side(side):
    """Convierte un Side con color theme/indexed a RGB explícito (negro)."""
    if not side or not side.style:
        return side
    try:
        ct = side.color.type if side.color else 'rgb'
        if ct in ('theme', 'indexed'):
            color = 'FF000000'
        elif ct == 'rgb':
            color = side.color.rgb if side.color.rgb != '00000000' else 'FF000000'
        else:
            color = 'FF000000'
    except:
        color = 'FF000000'
    return Side(style=side.style, color=color)

def normalize_border(border):
    """Normaliza todos los lados de un borde a colores RGB explícitos."""
    if not border:
        return border
    return Border(
        left=normalize_border_side(border.left),
        right=normalize_border_side(border.right),
        top=normalize_border_side(border.top),
        bottom=normalize_border_side(border.bottom),
        diagonal=border.diagonal,
        diagonal_direction=border.diagonal_direction,
    )


# ─────────────────────────────────────────────
#  HELPER: COPIAR HOJA ENTRE WORKBOOKS
# ─────────────────────────────────────────────
def copiar_hoja(ws_origen, wb_destino, nombre_hoja):
    """Copia una hoja completa (celdas, formato, merges, dimensiones) a otro workbook."""
    ws_dest = wb_destino.create_sheet(nombre_hoja)

    # Dimensiones de columnas y filas
    for col, cd in ws_origen.column_dimensions.items():
        ws_dest.column_dimensions[col].width = cd.width
        ws_dest.column_dimensions[col].hidden = cd.hidden
    for row, rd in ws_origen.row_dimensions.items():
        ws_dest.row_dimensions[row].height = rd.height
        ws_dest.row_dimensions[row].hidden = rd.hidden

    # Celdas combinadas (merged) — aplicar PRIMERO
    for rng in ws_origen.merged_cells.ranges:
        ws_dest.merge_cells(str(rng))

    # Copiar celdas con valor y estilo
    # IMPORTANTE: las celdas secundarias de rangos merged (MergedCell)
    # también tienen bordes — deben copiarse para que el formato sea correcto
    for row in ws_origen.iter_rows():
        for cell in row:
            dest_cell = ws_dest.cell(row=cell.row, column=cell.column)

            # Valor: solo funciona en celdas normales, no en MergedCell secundarias
            try:
                dest_cell.value = cell.value
            except AttributeError:
                pass  # MergedCell secundaria — sin valor

            # Estilos: intentar copiar en todas las celdas (incluidas MergedCell)
            if cell.has_style:
                # Borde: fundamental para celdas secundarias de merged ranges
                try:
                    dest_cell.border = normalize_border(cell.border)
                except Exception:
                    pass
                # Resto de estilos solo para celdas normales
                try:
                    dest_cell.font      = copy(cell.font)
                    dest_cell.alignment = copy(cell.alignment)
                    dest_cell.number_format = cell.number_format
                    rgb = get_fill_rgb(cell.fill)
                    if rgb:
                        dest_cell.fill = PatternFill(fill_type='solid', fgColor=rgb)
                    else:
                        dest_cell.fill = copy(cell.fill)
                except AttributeError:
                    pass  # MergedCell secundaria — solo border importa
    return ws_dest


def print_setup(ws, orientacion="portrait"):
    """Configura impresión A4, ajuste al ancho de la página, márgenes mínimos."""
    try:
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
        try:
            if ws.sheet_properties is None:
                ws.sheet_properties = WorksheetProperties()
            if ws.sheet_properties.pageSetUpPr is None:
                ws.sheet_properties.pageSetUpPr = PageSetupProperties()
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        except Exception:
            pass
        ws.page_setup.paperSize   = ws.PAPERSIZE_A4
        ws.page_setup.orientation = orientacion
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale       = 100
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered   = False
        ws.page_margins = PageMargins(
            left=0.2, right=0.2, top=0.3, bottom=0.3,
            header=0.0, footer=0.0
        )
    except Exception as e:
        print(f"⚠ print_setup error (no crítico): {e}")


def generar_ficha_completa(d):
    """Genera Excel con 4 hojas: 3 evaluación + certificado de aptitud."""
    import gc
    wb_final = Workbook()
    wb_final.remove(wb_final.active)

    for fn, nombre_hoja in [
        (llenar_hoja1,       "EVALUACIÓN OCUPACIONAL 1-3"),
        (llenar_hoja2,       "EVALUACIÓN OCUPACIONAL 2-3"),
        (llenar_hoja3,       "EVALUACIÓN OCUPACIONAL 3-3"),
        (llenar_certificado, "CERTIFICADO DE APTITUD LABORAL"),
    ]:
        buf = fn(d)
        wb_src = load_workbook(io.BytesIO(buf.read()))
        copiar_hoja(wb_src.active, wb_final, nombre_hoja)
        wb_src.close()
        del wb_src, buf
        gc.collect()

    out = io.BytesIO()
    wb_final.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────────
#  RUTAS API
# ─────────────────────────────────────────────
@app.route("/api/descargar-ficha", methods=["POST"])
def api_ficha():
    try:
        d = request.json
        nombre = f"{d.get('primerApellido','SN')}_{d.get('primerNombre','SN')}".replace(" ","_")
        fecha  = d.get("fechaAtencion","SFecha")
        buf = generar_ficha_completa(d)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"FichaOcupacional_{nombre}_{fecha}.xlsx"
        )
    except Exception as e:
        import traceback
        print("ERROR en /api/descargar-ficha:", traceback.format_exc())
        return f"Error: {str(e)}", 500

@app.route("/api/descargar-certificado", methods=["POST"])
def api_cert():
    try:
        d = request.json
        nombre = f"{d.get('primerApellido','SN')}_{d.get('primerNombre','SN')}".replace(" ","_")
        fecha  = d.get("fechaAtencion","SFecha")
        buf = llenar_certificado(d)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"Certificado_Aptitud_{nombre}_{fecha}.xlsx"
        )
    except Exception as e:
        import traceback
        print("ERROR en /api/descargar-certificado:", traceback.format_exc())
        return f"Error interno: {str(e)}", 500

@app.route("/api/descargar-historia", methods=["POST"])
def api_historia():
    d = request.json
    nombre = f"{d.get('primerApellido','SN')}_{d.get('primerNombre','SN')}".replace(" ","_")
    fecha  = d.get("fechaAtencion","SFecha")
    buf = generar_ficha_completa(d)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"FichaOcupacional_{nombre}_{fecha}.xlsx"
    )

@app.route("/api/descargar-consentimiento", methods=["POST"])
def api_consentimiento():
    from docx import Document
    from docx.shared import Pt
    d = request.json or {}
    nombre = f"{d.get('primerNombre','')} {d.get('primerApellido','')}".strip() or "Paciente"
    cedula = d.get("cedula","") or d.get("nroHistoria","")
    empresa = d.get("institucion","")
    fecha = d.get("fechaAtencion","")
    plantilla = os.path.join(PLANTILLAS, "CONSENTIMIENTO_INFORMADO.docx")
    if not os.path.exists(plantilla):
        return jsonify({"ok": False, "error": "Plantilla no encontrada"}), 404
    doc = Document(plantilla)
    reemplazos = {
        "{{NOMBRE}}": nombre,
        "{{CEDULA}}": cedula,
        "{{EMPRESA}}": empresa,
        "{{FECHA}}": fecha,
    }
    for p in doc.paragraphs:
        for run in p.runs:
            for k, v in reemplazos.items():
                if k in run.text:
                    run.text = run.text.replace(k, v)
    buf = io.BytesIO()
    doc.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        as_attachment=True,
        download_name=f"Consentimiento_{nombre}.docx"
    )

@app.route("/api/agregar-matriz", methods=["POST"])
def api_agregar():
    d = request.json
    nombre = f"{d.get('primerApellido','')} {d.get('segundoApellido','')} {d.get('primerNombre','')} {d.get('segundoNombre','')}".strip()
    APTITUD_TEXTO = {"APTO":"APTO","APTO_OBS":"APTO EN OBSERVACIÓN","APTO_LIM":"APTO CON LIMITACIONES","NO_APTO":"NO APTO"}
    fila = {
        "N°": db_contar() + 1,
        "NOMBRE COMPLETO": nombre,
        "N.° de historia clínica": d.get("nroHistoria",""),
        "Empresa / institución": d.get("institucion",""),
        "RUC": d.get("ruc",""),
        "ESTABLECIMIENTO": d.get("establecimiento",""),
        "Puesto / cargo (CIUO)": d.get("puestoCIUO",""),
        "SEXO": d.get("sexo",""),
        "FECHA NACIMIENTO": d.get("fechaNacimiento",""),
        "EDAD (años)": d.get("edad",""),
        "GRUPO SANGUÍNEO": d.get("grupoSanguineo",""),
        "GRUPO AT. PRIORITARIA": ", ".join(filter(None,[
            "Embarazada" if d.get("embarazada") else "",
            "Discapacidad" if d.get("discapacidad") else "",
            "Enf.Catastrófica" if d.get("enfCatastrofica") else "",
            "Adulto Mayor" if d.get("adultoMayor") else "",
        ])) or "Ninguno",
        "Tipo de evaluación": d.get("tipoEvaluacion",""),
        "Fecha de atención": d.get("fechaAtencion",""),
        "Fecha de ingreso": d.get("fechaIngreso",""),
        "Temperatura (°C)": d.get("temperatura",""),
        "Presión arterial": d.get("pa",""),
        "FC (lat/min)": d.get("fc",""),
        "FR (fr/min)": d.get("fr",""),
        "SpO₂ (%)": d.get("spo2",""),
        "Peso (kg)": d.get("peso",""),
        "Talla (cm)": d.get("talla",""),
        "IMC (kg/m²)": d.get("imc",""),
        "Perímetro abdominal (cm)": d.get("perAbdominal",""),
        "Aptitud": APTITUD_TEXTO.get(d.get("aptitud",""),""),
        "Observaciones de aptitud": d.get("aptitudObs",""),
        "DIAGNÓSTICO 1 CIE-10": (d.get("diagnosticos",[{}])[0] or {}).get("cie10",""),
        "DESCRIPCIÓN DX1": (d.get("diagnosticos",[{}])[0] or {}).get("desc",""),
        "DIAGNÓSTICO 2 CIE-10": (d.get("diagnosticos",[{},{}])[1] if len(d.get("diagnosticos",[])) > 1 else {}).get("cie10",""),
        "DESCRIPCIÓN DX2": (d.get("diagnosticos",[{},{}])[1] if len(d.get("diagnosticos",[])) > 1 else {}).get("desc",""),
        "Recomendaciones": d.get("recomendaciones",""),
        "Médico responsable": d.get("nombreProf",""),
        "Código médico": d.get("codigoMedico",""),
    }
    try:
        db_insertar(fila)
        total = db_contar()
        return jsonify({"ok": True, "total": total, "nombre": nombre})
    except Exception as e:
        # Si Supabase falla, guardar en memoria local como respaldo
        _matriz_local.append(fila)
        total = len(_matriz_local)
        print(f"⚠ Supabase error, guardado local: {e}")
        return jsonify({"ok": True, "total": total, "nombre": nombre})

@app.route("/api/descargar-matriz", methods=["GET"])
def api_matriz():
    try:
        filas = db_listar()
    except Exception as e:
        filas = []
    # Si Supabase está vacío pero hay datos locales, usar locales
    if not filas and _matriz_local:
        filas = _matriz_local
    if not filas:
        return jsonify({"error": "Matriz vacía"}), 400
    wb = Workbook()
    ws = wb.active
    ws.title = "MATRIZ DE SEGUIMIENTO"
    headers = list(filas[0].keys())
    fill_hdr = PatternFill("solid", fgColor="FF1556A0")
    font_hdr = Font(bold=True, color="FFFFFFFF", size=9)
    thin = Side(style='thin', color='FF000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    fill_even = PatternFill("solid", fgColor="FFE8EEF8")
    font_data = Font(size=9)
    for row_i, fila in enumerate(filas, 2):
        fill = fill_even if row_i % 2 == 0 else PatternFill("solid", fgColor="FFFFFFFF")
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=row_i, column=col_i, value=fila.get(h,""))
            cell.fill = fill
            cell.font = font_data
            cell.alignment = Alignment(wrap_text=True)
            cell.border = border
    widths = {"N°":5,"NOMBRE COMPLETO":30,"N.° de historia clínica":15,"Empresa / institución":25,"RUC":14,
              "ESTABLECIMIENTO":22,"Puesto / cargo (CIUO)":22,"SEXO":8,"FECHA NACIMIENTO":14,"EDAD (años)":8,
              "GRUPO SANGUÍNEO":10,"GRUPO AT. PRIORITARIA":20,"Tipo de evaluación":14,"Fecha de atención":14,
              "Fecha de ingreso":14,"Temperatura (°C)":10,"Presión arterial":10,"FC (lat/min)":8,"FR (fr/min)":8,
              "SpO₂ (%)":8,"Peso (kg)":8,"Talla (cm)":8,"IMC (kg/m²)":10,"Perímetro abdominal (cm)":14,
              "Aptitud":22,"Observaciones de aptitud":30,"DIAGNÓSTICO 1 CIE-10":12,"DESCRIPCIÓN DX1":30,
              "DIAGNÓSTICO 2 CIE-10":12,"DESCRIPCIÓN DX2":30,"Recomendaciones":40,
              "Médico responsable":25,"Código médico":12}
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Matriz_Fichas_Medicas_{date.today().isoformat()}.xlsx")

@app.route("/api/limpiar-matriz", methods=["POST"])
def api_limpiar():
    n = db_contar()
    db_limpiar()
    return jsonify({"ok": True, "eliminados": n})

@app.route("/api/conteo-matriz")
def api_conteo():
    filas = db_listar()
    return jsonify({"total": len(filas),
                    "nombres": [f"{m.get('N°',i+1)}. {m.get('NOMBRE COMPLETO','')}" for i,m in enumerate(filas)]})

@app.route("/api/estadisticas")
def api_estadisticas():
    filas = db_listar()
    total = len(filas)
    if total == 0:
        return jsonify({"total": 0})

    from collections import Counter, defaultdict

    # ── Edad ──────────────────────────────────────
    rangos_edad = {"18-25":0,"26-35":0,"36-45":0,"46-55":0,"56-65":0,">65":0,"N/D":0}
    for f in filas:
        try:
            e = int(float(str(f.get("EDAD (años)","") or f.get("edad","") or 0)))
            if   e <= 25: rangos_edad["18-25"] += 1
            elif e <= 35: rangos_edad["26-35"] += 1
            elif e <= 45: rangos_edad["36-45"] += 1
            elif e <= 55: rangos_edad["46-55"] += 1
            elif e <= 65: rangos_edad["56-65"] += 1
            elif e > 65:  rangos_edad[">65"]   += 1
            else:         rangos_edad["N/D"]   += 1
        except: rangos_edad["N/D"] += 1

    # ── Género ────────────────────────────────────
    genero = Counter()
    for f in filas:
        g = (f.get("SEXO","") or f.get("sexo","") or "N/D").strip()
        genero[g if g else "N/D"] += 1

    # ── Grupo sanguíneo ───────────────────────────
    gs = Counter()
    for f in filas:
        v = (f.get("GRUPO SANGUÍNEO","") or f.get("grupoSanguineo","") or "N/D").strip()
        gs[v if v else "N/D"] += 1

    # ── Aptitud ───────────────────────────────────
    AMAP = {"APTO":"Apto","APTO_OBS":"Apto en observación",
            "APTO_LIM":"Apto con limitaciones","NO_APTO":"No apto"}
    aptitud_cnt = Counter()
    for f in filas:
        v = f.get("APTITUD","") or f.get("aptitud","") or ""
        aptitud_cnt[AMAP.get(v, v) if v else "N/D"] += 1

    # ── IMC ───────────────────────────────────────
    imc_cnt = {"Bajo peso":0,"Normal":0,"Sobrepeso":0,
               "Obesidad I":0,"Obesidad II":0,"Obesidad III":0,"N/D":0}
    for f in filas:
        try:
            imc = float(str(f.get("IMC (kg/m²)","") or f.get("imc","") or 0).replace(",","."))
            if   imc <= 0:    imc_cnt["N/D"] += 1
            elif imc < 18.5:  imc_cnt["Bajo peso"] += 1
            elif imc < 25:    imc_cnt["Normal"] += 1
            elif imc < 30:    imc_cnt["Sobrepeso"] += 1
            elif imc < 35:    imc_cnt["Obesidad I"] += 1
            elif imc < 40:    imc_cnt["Obesidad II"] += 1
            else:             imc_cnt["Obesidad III"] += 1
        except: imc_cnt["N/D"] += 1

    # ── Diagnósticos más frecuentes ───────────────
    dx_cnt = Counter()
    for f in filas:
        dx1_cie  = f.get("DIAGNÓSTICO 1 CIE-10","") or ""
        dx1_desc = f.get("DESCRIPCIÓN DX1","") or ""
        dx2_cie  = f.get("DIAGNÓSTICO 2 CIE-10","") or ""
        dx2_desc = f.get("DESCRIPCIÓN DX2","") or ""
        if dx1_cie.strip(): dx_cnt[f"{dx1_cie.strip()} — {dx1_desc.strip()}"] += 1
        if dx2_cie.strip(): dx_cnt[f"{dx2_cie.strip()} — {dx2_desc.strip()}"] += 1
    top_dx = [{"dx": k, "n": v} for k, v in dx_cnt.most_common(10)]

    # ── Grupos prioritarios ───────────────────────
    adultos_mayores = sum(1 for f in filas if ">65" == (
        ">65" if (lambda e: e>65 if e else False)(
            __builtins__["int"](float(str(f.get("EDAD (años)","") or 0))) if str(f.get("EDAD (años)","") or 0).replace(".","").isdigit() else 0
        ) else ""))

    gp = {
        "Adultos mayores (>65 años)": rangos_edad[">65"],
        "Embarazadas":    sum(1 for f in filas if f.get("GRUPO AT. PRIORITARIA","") and "Embarazada" in str(f.get("GRUPO AT. PRIORITARIA",""))),
        "Discapacidad":   sum(1 for f in filas if f.get("GRUPO AT. PRIORITARIA","") and "Discapacidad" in str(f.get("GRUPO AT. PRIORITARIA",""))),
        "Enf. Catastrófica": sum(1 for f in filas if f.get("GRUPO AT. PRIORITARIA","") and "Catastrófica" in str(f.get("GRUPO AT. PRIORITARIA",""))),
    }

    # ── ECNT (Enfermedades Crónicas No Transmisibles) ──
    ECNT_CODIGOS = {
        "E10","E11","E12","E13","E14",  # Diabetes
        "I10","I11","I12","I13","I15",  # Hipertensión
        "E66",                           # Obesidad
        "E78",                           # Dislipidemia
        "J44","J45",                     # EPOC/Asma
        "I20","I21","I25",              # Cardiopatía
        "K70","K73","K74",              # Enf. hepática
    }
    ecnt = 0
    for f in filas:
        for campo in ["DIAGNÓSTICO 1 CIE-10","DIAGNÓSTICO 2 CIE-10"]:
            cie = str(f.get(campo,"") or "").strip()[:3]
            if cie in ECNT_CODIGOS:
                ecnt += 1
                break

    # ── Cargos ────────────────────────────────────
    cargos_cnt = Counter()
    for f in filas:
        c = (f.get("PUESTO/CARGO (CIUO)","") or f.get("puestoCIUO","") or "N/D").strip()
        cargos_cnt[c if c else "N/D"] += 1
    top_cargos = [{"cargo": k, "n": v} for k, v in cargos_cnt.most_common(15)]

    return jsonify({
        "total": total,
        "edad": rangos_edad,
        "genero": dict(genero),
        "grupo_sanguineo": dict(gs),
        "aptitud": dict(aptitud_cnt),
        "imc": imc_cnt,
        "diagnosticos": top_dx,
        "grupos_prioritarios": gp,
        "ecnt": ecnt,
        "cargos": top_cargos,
        "total_cargos": len(cargos_cnt),
    })


    html_path = os.path.join(BASE, "index.html")
    with open(html_path, encoding="utf-8") as f:
        return f.read()

# ─────────────────────────────────────────────
#  BORRADORES — guardar / cargar / eliminar
# ─────────────────────────────────────────────
@app.route("/api/borrador/guardar", methods=["POST"])
def borrador_guardar():
    if not supa:
        return jsonify({"ok": False, "error": "Supabase no configurado"}), 503
    d = request.json or {}
    cedula = (d.get("cedula") or d.get("nroHistoria") or "").strip()
    if not cedula:
        return jsonify({"ok": False, "error": "Cédula requerida"}), 400
    try:
        # Upsert: si ya existe ese borrador lo sobreescribe
        supa.table("borradores_femo").upsert(
            {"cedula": cedula, "data": d},
            on_conflict="cedula"
        ).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/borrador/<cedula>", methods=["GET"])
def borrador_cargar(cedula):
    if not supa:
        return jsonify({"ok": False, "error": "Supabase no configurado"}), 503
    try:
        r = supa.table("borradores_femo").select("data").eq("cedula", cedula).execute()
        if r.data:
            return jsonify({"ok": True, "data": r.data[0]["data"]})
        return jsonify({"ok": False, "error": "no_encontrado"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/borrador/<cedula>", methods=["DELETE"])
def borrador_eliminar(cedula):
    if not supa:
        return jsonify({"ok": False, "error": "Supabase no configurado"}), 503
    try:
        supa.table("borradores_femo").delete().eq("cedula", cedula).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/")
def index():
    html_path = os.path.join(BASE, "index.html")
    with open(html_path, encoding="utf-8") as f:
        return f.read()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    print("=" * 55)
    print("  FICHA MÉDICA OCUPACIONAL – CASISO S.A.S.")
    print(f"  Servidor en puerto {port}")
    print("=" * 55)
    app.run(host="0.0.0.0", port=port, debug=debug)


# ═══════════════════════════════════════════════════════════════
#  DASHBOARD — CRUD DE REGISTROS Y ESTADÍSTICAS AMPLIADAS
# ═══════════════════════════════════════════════════════════════

# Mapeo CIE-10 (3 caracteres) → Sistema orgánico
_CIE_SISTEMA = {
    "E66":"Metabólico","E67":"Metabólico","E68":"Metabólico",
    "E78":"Cardiovascular/Lipídico","E79":"Metabólico",
    "R74":"Hepático","K70":"Hepático","K73":"Hepático","K74":"Hepático",
    "D64":"Hematológico","D50":"Hematológico","D51":"Hematológico","D53":"Hematológico",
    "J02":"Respiratorio/ORL","J00":"Respiratorio/ORL","J03":"Respiratorio/ORL","J06":"Respiratorio/ORL",
    "H90":"Auditivo","H91":"Auditivo","H83":"Auditivo",
    "B82":"Digestivo","B80":"Digestivo","K63":"Digestivo",
    "M41":"Osteomuscular","M54":"Osteomuscular","M47":"Osteomuscular","M51":"Osteomuscular",
    "R73":"Metabólico","E10":"Metabólico","E11":"Metabólico","E14":"Metabólico",
    "I10":"Cardiovascular/Lipídico","I11":"Cardiovascular/Lipídico","I15":"Cardiovascular/Lipídico",
    "N39":"Urológico","N30":"Urológico","R82":"Urológico",
    "E03":"Metabólico","E01":"Metabólico",
}

# Mapeo CIE-10 → ECNT específica
_CIE_ECNT = {
    "E78":"Dislipidemia",
    "E66":"Obesidad",
    "E79":"Hiperuricemia",
    "R74":"Enzimas hepáticas elevadas",
    "D64":"Anemia crónica","D50":"Anemia crónica","D51":"Anemia crónica",
    "R73":"Prediabetes/Hiperglicemia",
    "E10":"Diabetes","E11":"Diabetes","E14":"Diabetes",
    "I10":"HTA probable","I11":"HTA probable",
    "E03":"Hipotiroidismo",
    "M41":"Escoliosis",
    "H90":"Hipoacusia",
    "B82":"Parasitosis","B80":"Parasitosis",
}


def db_listar_con_id():
    """Retorna registros incluyendo el ID de fila de Supabase para CRUD."""
    if supa:
        try:
            res = supa.table(TABLE_NAME).select("id, data").order("id").execute()
            result = []
            for r in res.data:
                item = dict(r["data"])
                item["_id"] = r["id"]
                result.append(item)
            return result
        except Exception as e:
            print(f"⚠ db_listar_con_id error: {e}")
    return [dict(f, _id=i+1) for i, f in enumerate(_matriz_local)]


@app.route("/api/matriz/registros")
def api_registros():
    return jsonify(db_listar_con_id())


@app.route("/api/matriz/editar", methods=["POST"])
def api_editar_registro():
    d = request.json or {}
    row_id = d.pop("_id", None)
    if not row_id:
        return jsonify({"ok": False, "error": "ID requerido"}), 400
    if supa:
        try:
            supa.table(TABLE_NAME).update({"data": d}).eq("id", int(row_id)).execute()
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500
    for i, f in enumerate(_matriz_local):
        if str(f.get("N°")) == str(d.get("N°")):
            _matriz_local[i] = d
            return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "No encontrado"}), 404


@app.route("/api/matriz/eliminar/<int:row_id>", methods=["DELETE"])
def api_eliminar_registro(row_id):
    global _matriz_local
    if supa:
        try:
            supa.table(TABLE_NAME).delete().eq("id", row_id).execute()
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500
    _matriz_local = [f for i, f in enumerate(_matriz_local) if i+1 != row_id]
    return jsonify({"ok": True})


@app.route("/api/estadisticas2")
def api_estadisticas2():
    """Estadísticas ampliadas con sistemas, ECNT y datos CRUD."""
    from collections import Counter
    filas = db_listar()
    total = len(filas)
    if total == 0:
        return jsonify({"total": 0})

    # ── Edad ──
    rangos = {"18-25":0,"26-35":0,"36-45":0,"46-55":0,"56-65":0,">65":0,"N/D":0}
    for f in filas:
        try:
            e = int(float(str(f.get("EDAD (años)","") or 0)))
            if   e <= 25: rangos["18-25"] += 1
            elif e <= 35: rangos["26-35"] += 1
            elif e <= 45: rangos["36-45"] += 1
            elif e <= 55: rangos["46-55"] += 1
            elif e <= 65: rangos["56-65"] += 1
            elif e > 65:  rangos[">65"]   += 1
            else:         rangos["N/D"]   += 1
        except: rangos["N/D"] += 1

    # ── Género ──
    genero = Counter()
    for f in filas:
        g = (f.get("SEXO","") or "N/D").strip()
        genero[g or "N/D"] += 1

    # ── Grupo sanguíneo ──
    gs = Counter()
    for f in filas:
        v = (f.get("GRUPO SANGUÍNEO","") or "N/D").strip()
        gs[v or "N/D"] += 1

    # ── Aptitud ──
    AMAP = {"APTO":"Apto","APTO EN OBSERVACIÓN":"Apto en observación",
            "APTO_OBS":"Apto en observación","APTO_LIM":"Apto con limitaciones",
            "APTO CON LIMITACIONES":"Apto con limitaciones",
            "NO APTO":"No apto","NO_APTO":"No apto"}
    apt_cnt = Counter()
    for f in filas:
        v = (f.get("Aptitud","") or f.get("APTITUD","") or "").strip().upper()
        apt_cnt[AMAP.get(v, v.title()) if v else "N/D"] += 1

    # ── IMC ──
    imc_cnt = {"Bajo peso":0,"Normal":0,"Sobrepeso":0,
               "Obesidad I":0,"Obesidad II":0,"Obesidad III":0,"N/D":0}
    for f in filas:
        try:
            imc = float(str(f.get("IMC (kg/m²)","") or 0).replace(",","."))
            if   imc <= 0:    imc_cnt["N/D"] += 1
            elif imc < 18.5:  imc_cnt["Bajo peso"] += 1
            elif imc < 25:    imc_cnt["Normal"] += 1
            elif imc < 30:    imc_cnt["Sobrepeso"] += 1
            elif imc < 35:    imc_cnt["Obesidad I"] += 1
            elif imc < 40:    imc_cnt["Obesidad II"] += 1
            else:             imc_cnt["Obesidad III"] += 1
        except: imc_cnt["N/D"] += 1

    # ── Diagnósticos ──
    dx_cnt = Counter()
    for f in filas:
        for c1, c2 in [("DIAGNÓSTICO 1 CIE-10","DESCRIPCIÓN DX1"),("DIAGNÓSTICO 2 CIE-10","DESCRIPCIÓN DX2")]:
            cie = (f.get(c1,"") or "").strip()
            desc = (f.get(c2,"") or "").strip()
            if cie:
                key = f"{cie} — {desc}" if desc else cie
                dx_cnt[key] += 1
    top_dx = [{"dx":k,"n":v} for k,v in dx_cnt.most_common(12)]

    # ── Sistemas afectados ──
    sis_cnt = Counter()
    for f in filas:
        seen = set()
        for campo in ["DIAGNÓSTICO 1 CIE-10","DIAGNÓSTICO 2 CIE-10"]:
            cie = str(f.get(campo,"") or "").strip()[:3]
            sistema = _CIE_SISTEMA.get(cie)
            if sistema and sistema not in seen:
                sis_cnt[sistema] += 1
                seen.add(sistema)

    # ── ECNT detalle ──
    ecnt_cnt = Counter()
    ecnt_pacs = set()
    for i, f in enumerate(filas):
        for campo in ["DIAGNÓSTICO 1 CIE-10","DIAGNÓSTICO 2 CIE-10"]:
            cie = str(f.get(campo,"") or "").strip()[:3]
            ecnt_name = _CIE_ECNT.get(cie)
            if ecnt_name:
                ecnt_cnt[ecnt_name] += 1
                ecnt_pacs.add(i)

    # ── Grupos prioritarios ──
    gp = {
        "Adultos mayores >65": rangos[">65"],
        "Embarazadas":   sum(1 for f in filas if "Embarazada" in str(f.get("GRUPO AT. PRIORITARIA",""))),
        "Discapacidad":  sum(1 for f in filas if "Discapacidad" in str(f.get("GRUPO AT. PRIORITARIA",""))),
        "Enf. Catastrófica": sum(1 for f in filas if "Catastrófica" in str(f.get("GRUPO AT. PRIORITARIA",""))),
    }

    # ── Cargos ──
    cargos_cnt = Counter()
    for f in filas:
        c = (f.get("Puesto / cargo (CIUO)","") or f.get("puestoCIUO","") or "N/D").strip()
        cargos_cnt[c or "N/D"] += 1
    top_cargos = [{"cargo":k,"n":v} for k,v in cargos_cnt.most_common(15)]

    # ── Promedio de edad ──
    edades = []
    for f in filas:
        try:
            e = int(float(str(f.get("EDAD (años)","") or 0)))
            if e > 0: edades.append(e)
        except: pass
    avg_edad = round(sum(edades)/len(edades)) if edades else 0

    return jsonify({
        "total": total,
        "edad": rangos,
        "genero": dict(genero),
        "grupo_sanguineo": dict(gs),
        "aptitud": dict(apt_cnt),
        "imc": imc_cnt,
        "diagnosticos": top_dx,
        "sistemas": dict(sis_cnt),
        "ecnt_detalle": dict(ecnt_cnt),
        "ecnt": len(ecnt_pacs),
        "grupos_prioritarios": gp,
        "cargos": top_cargos,
        "total_cargos": len(cargos_cnt),
        "avg_edad": avg_edad,
    })
